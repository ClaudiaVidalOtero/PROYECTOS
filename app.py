from datetime import datetime, timedelta
import json
import os
from pathlib import Path

from flask import Flask, flash, jsonify, redirect, render_template, request, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from dateutil.relativedelta import relativedelta
from unidecode import unidecode
import datetime

app = Flask(__name__)
basedir = Path(__file__).parent
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL",
    f"sqlite:///{basedir / 'competitions.db'}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-key-change-in-production")

db = SQLAlchemy(app)


# Models
class Competition(db.Model):
    __tablename__ = "competitions"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(255), nullable=False)
    event_date = db.Column(db.Date, nullable=False, default=datetime.date.today)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    
    categories = db.relationship("Category", back_populates="competition", cascade="all, delete-orphan")
    participants = db.relationship("Participant", back_populates="competition", cascade="all, delete-orphan")
    matches = db.relationship("Match", back_populates="competition", cascade="all, delete-orphan")


class Category(db.Model):
    __tablename__ = "categories"
    id = db.Column(db.Integer, primary_key=True)
    competition_id = db.Column(db.Integer, db.ForeignKey("competitions.id"), nullable=False)
    name = db.Column(db.String(255), nullable=False)
    modality = db.Column(db.String(10), nullable=False)  # "kata" or "kumite"
    group = db.Column(db.String(100), nullable=True)  # Group/Subcategory for organization
    criteria = db.Column(db.JSON, nullable=False)  # e.g., {"age_min": 10, "age_max": 12, "gender": "M", "grade": "3"}
    
    competition = db.relationship("Competition", back_populates="categories")
    participant_links = db.relationship("ParticipantCategory", back_populates="category", cascade="all, delete-orphan")
    participants = db.relationship(
        "Participant",
        secondary="participant_categories",
        back_populates="categories",
        viewonly=True,
        overlaps="participant_links"
    )
    matches = db.relationship("Match", back_populates="category", cascade="all, delete-orphan")


class ParticipantCategory(db.Model):
    __tablename__ = "participant_categories"
    participant_id = db.Column(db.Integer, db.ForeignKey("participants.id"), primary_key=True)
    category_id = db.Column(db.Integer, db.ForeignKey("categories.id"), primary_key=True)
    score = db.Column(db.Float, nullable=True)
    tiebreaker_round = db.Column(db.Integer, nullable=True)

    participant = db.relationship("Participant", back_populates="category_links", overlaps="participants")
    category = db.relationship("Category", back_populates="participant_links", overlaps="participants")


class Participant(db.Model):
    __tablename__ = "participants"
    id = db.Column(db.Integer, primary_key=True)
    competition_id = db.Column(db.Integer, db.ForeignKey("competitions.id"), nullable=False)
    
    name = db.Column(db.String(255), nullable=False)
    birthdate = db.Column(db.Date, nullable=False)
    gender = db.Column(db.String(1), nullable=False)  # M, F, X
    grade = db.Column(db.String(20), nullable=True)  # "3", "1 dan", etc.
    weight = db.Column(db.Float, nullable=True)
    
    kata_participation = db.Column(db.Boolean, default=False)
    kumite_participation = db.Column(db.Boolean, default=False)
    
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    
    competition = db.relationship("Competition", back_populates="participants")
    category_links = db.relationship(
        "ParticipantCategory",
        back_populates="participant",
        cascade="all, delete-orphan",
        overlaps="participants"
    )
    categories = db.relationship(
        "Category",
        secondary="participant_categories",
        back_populates="participants",
        viewonly=True,
        overlaps="category_links"
    )
    
    def age_at_date(self, event_date):
        if not self.birthdate:
            return None
        age = relativedelta(event_date, self.birthdate)
        return age.years


class Match(db.Model):
    __tablename__ = "matches"
    id = db.Column(db.Integer, primary_key=True)
    competition_id = db.Column(db.Integer, db.ForeignKey("competitions.id"), nullable=False)
    category_id = db.Column(db.Integer, db.ForeignKey("categories.id"), nullable=False)
    
    participant1_id = db.Column(db.Integer, db.ForeignKey("participants.id"), nullable=False)
    participant2_id = db.Column(db.Integer, db.ForeignKey("participants.id"), nullable=True)
    
    round_number = db.Column(db.Integer, nullable=False, default=1)
    bracket_type = db.Column(db.String(20), default="elimination")  # elimination, roundrobin, bye
    
    winner_id = db.Column(db.Integer, db.ForeignKey("participants.id"), nullable=True)
    status = db.Column(db.String(20), default="pending")  # pending, completed, bye
    
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.datetime.utcnow)
    
    competition = db.relationship("Competition", back_populates="matches")
    category = db.relationship("Category", back_populates="matches")
    participant1 = db.relationship("Participant", foreign_keys=[participant1_id])
    participant2 = db.relationship("Participant", foreign_keys=[participant2_id])
    winner = db.relationship("Participant", foreign_keys=[winner_id])


def calculate_age_at_event(birthdate, event_date):
    """Calculate exact age on event date."""
    if not birthdate or not event_date:
        return None
    age = relativedelta(event_date, birthdate)
    return age.years


def parse_grade(value):
    """Parse grade from string like '3' or '1 dan'."""
    if not value:
        return None
    text = str(value).strip().lower()
    return text if text else None


def grade_to_numeric(grade_str):
    """
    Convert grade string to numeric value for comparison.
    12 kyu = 0, 11 kyu = 1, ..., 1 kyu = 11, 1 dan = 12, 2 dan = 13, ..., 5 dan = 16
    Higher number = higher rank
    """
    if not grade_str:
        return None
    
    grade_str = str(grade_str).strip().lower()
    
    # Check if it's a dan rank
    if "dan" in grade_str:
        try:
            dan_num = int(grade_str.split("dan")[0].strip())
            return 11 + dan_num  # 1 dan = 12, 2 dan = 13, etc.
        except:
            return None
    
    # Check if it's a kyu rank
    try:
        # Handle decimal numbers like "7.0"
        if "kyu" in grade_str:
            kyu_num = float(grade_str.split("kyu")[0].strip())
        else:
            kyu_num = float(grade_str.strip())
        kyu_num = int(kyu_num)  # Convert to int
        # 12 kyu = 0, 11 kyu = 1, ..., 1 kyu = 11
        return max(0, 12 - kyu_num)
    except:
        return None


def normalize_excel_columns(headers):
    """Normalize Excel column names to expected fields."""
    column_mapping = {
        "numero_participante": "numero",
        "nombre_participante": "name",
        "edad": "age",
        "fecha_nacimiento": "birthdate",
        "grado": "grade",
        "kata": "kata",
        "kumite": "kumite",
        "categoría": "category",
        "pagó": "paid",
        "peso": "weight",
        "género": "gender",
        "genero": "gender",
    }

    normalized = {}
    for i, header in enumerate(headers):
        header_clean = str(header).strip().lower() if header else ""
        mapped_name = column_mapping.get(header_clean, header_clean)
        normalized[mapped_name] = i

    return normalized


def load_participants_from_excel(file, competition):
    """Load participants from uploaded Excel file using openpyxl directly."""
    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active

        # Get headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        column_map = normalize_excel_columns(headers)

        # Debug: print available columns
        available_cols = list(column_map.keys())
        print(f"Available columns in Excel: {available_cols}")

        participants = []
        errors = []

        # Process each row starting from row 2
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Convert row tuple to dict using column mapping
                row_data = {}
                for col_name, col_idx in column_map.items():
                    if col_idx < len(row):
                        row_data[col_name] = row[col_idx]

                name = str(row_data.get("name", "")).strip()
                if not name or name.lower() in ("none", "nan", ""):
                    errors.append(f"Fila {row_idx}: falta nombre")
                    continue

                birthdate_val = row_data.get("birthdate")
                if not birthdate_val or str(birthdate_val).lower() in ("none", "nan", ""):
                    errors.append(f"Fila {row_idx} ({name}): falta fecha de nacimiento")
                    continue

                # Parse birthdate - try multiple formats
                birthdate = None
                birthdate_str = str(birthdate_val).strip()

                # Try DD/MM/YYYY first
                try:
                    birthdate = datetime.datetime.strptime(birthdate_str, "%d/%m/%Y").date()
                except ValueError:
                    # Try other common formats
                    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"]:
                        try:
                            birthdate = datetime.datetime.strptime(birthdate_str, fmt).date()
                            break
                        except ValueError:
                            continue

                if not birthdate:
                    errors.append(f"Fila {row_idx} ({name}): formato de fecha inválido '{birthdate_str}'")
                    continue

                gender = str(row_data.get("gender", "X")).strip().upper()
                if gender.startswith("M"):
                    gender = "M"
                elif gender.startswith("F"):
                    gender = "F"
                else:
                    gender = "X"

                kata = "X" in str(row_data.get("kata", "")).upper().strip()
                kumite = "X" in str(row_data.get("kumite", "")).upper().strip()

                grade = parse_grade(row_data.get("grade"))

                weight = None
                try:
                    weight_val = row_data.get("weight")
                    if weight_val is not None and str(weight_val).strip():
                        weight = float(str(weight_val).replace(",", "."))
                except (ValueError, TypeError):
                    pass

                participant = Participant(
                    competition_id=competition.id,
                    name=name,
                    birthdate=birthdate,
                    gender=gender,
                    grade=grade,
                    weight=weight,
                    kata_participation=kata,
                    kumite_participation=kumite,
                )

                # Check for duplicate by name in this competition
                existing = Participant.query.filter_by(competition_id=competition.id, name=name).first()
                if existing:
                    errors.append(f"Fila {row_idx} ({name}): participante ya existe")
                    continue

                participants.append(participant)

            except Exception as e:
                errors.append(f"Fila {row_idx}: {str(e)}")

        if participants:
            db.session.add_all(participants)
            db.session.commit()

        msg = f"Se cargaron {len(participants)} participantes."
        if errors:
            msg += f" Errores: {'; '.join(errors[:3])}"  # Show first 3 errors

        return len(participants) > 0, msg

    except Exception as e:
        return False, f"Error procesando Excel: {str(e)}"


def assign_participants_to_categories(competition, modality):
    """Re-assign ALL participants to categories based on modality."""
    if modality == "kata":
        all_participants = Participant.query.filter_by(
            competition_id=competition.id,
            kata_participation=True
        ).all()
    else:
        all_participants = Participant.query.filter_by(
            competition_id=competition.id,
            kumite_participation=True
        ).all()
    
    missing_data_category = Category.query.filter_by(
        competition_id=competition.id,
        name="faltan datos",
        modality=modality
    ).first()
    if not missing_data_category:
        missing_data_category = Category(
            competition_id=competition.id,
            name="faltan datos",
            modality=modality,
            criteria={}
        )
        db.session.add(missing_data_category)
        db.session.flush()
    
    extras_category = Category.query.filter_by(
        competition_id=competition.id,
        name="extras",
        modality=modality
    ).first()
    if not extras_category:
        extras_category = Category(
            competition_id=competition.id,
            name="extras",
            modality=modality,
            criteria={}
        )
        db.session.add(extras_category)
        db.session.flush()
    
    for participant in all_participants:
        # Remove existing assignments for this modality before reassigning
        for link in list(participant.category_links):
            if link.category and link.category.modality == modality:
                db.session.delete(link)
        
        age = calculate_age_at_event(participant.birthdate, competition.event_date)
        
        if modality == "kata":
            if age is None or participant.gender == "X" or not participant.grade:
                db.session.add(ParticipantCategory(participant=participant, category=missing_data_category))
                continue
        else:  # kumite
            if age is None or participant.gender == "X" or participant.weight is None:
                db.session.add(ParticipantCategory(participant=participant, category=missing_data_category))
                continue
        
        matching_category = None
        for cat in competition.categories:
            if cat.modality != modality or cat.name in ["faltan datos", "extras"]:
                continue
            
            criteria = cat.criteria or {}
            
            if "age_min" in criteria and age < criteria["age_min"]:
                continue
            if "age_max" in criteria and age > criteria["age_max"]:
                continue
            
            if "gender" in criteria and criteria["gender"] != "X" and criteria["gender"] != participant.gender:
                continue
            
            if modality == "kata":
                if participant.grade:
                    grade_numeric = grade_to_numeric(participant.grade)
                    if grade_numeric is not None:
                        if "grade_min" in criteria and grade_numeric < criteria["grade_min"]:
                            continue
                        if "grade_max" in criteria and grade_numeric > criteria["grade_max"]:
                            continue
                    else:
                        continue
                else:
                    continue
            
            if modality == "kumite":
                if "weight_min" in criteria and participant.weight < criteria["weight_min"]:
                    continue
                if "weight_max" in criteria and participant.weight > criteria["weight_max"]:
                    continue
            
            matching_category = cat
            break
        
        if matching_category:
            db.session.add(ParticipantCategory(participant=participant, category=matching_category))
        else:
            db.session.add(ParticipantCategory(participant=participant, category=extras_category))
    
    db.session.commit()


def generate_brackets(category):
    """Generate bracket matches for a category (elimination + round robin)."""
    participants = category.participants
    
    if len(participants) < 3:
        # Just create matches for all vs all if < 3
        for i, p1 in enumerate(participants):
            for p2 in participants[i+1:]:
                match = Match(
                    competition_id=category.competition_id,
                    category_id=category.id,
                    participant1_id=p1.id,
                    participant2_id=p2.id,
                    bracket_type="roundrobin",
                    status="pending"
                )
                db.session.add(match)
    elif len(participants) == 3:
        # Round robin
        for i, p1 in enumerate(participants):
            for p2 in participants[i+1:]:
                match = Match(
                    competition_id=category.competition_id,
                    category_id=category.id,
                    participant1_id=p1.id,
                    participant2_id=p2.id,
                    bracket_type="roundrobin",
                    status="pending"
                )
                db.session.add(match)
    else:
        # Elimination
        participants_copy = list(participants)
        round_num = 1
        while len(participants_copy) > 1:
            for i in range(0, len(participants_copy)-1, 2):
                p1 = participants_copy[i]
                p2 = participants_copy[i+1] if i+1 < len(participants_copy) else None
                
                if p2:
                    match = Match(
                        competition_id=category.competition_id,
                        category_id=category.id,
                        participant1_id=p1.id,
                        participant2_id=p2.id,
                        bracket_type="elimination",
                        round_number=round_num,
                        status="pending"
                    )
                    db.session.add(match)
                else:
                    # Bye
                    match = Match(
                        competition_id=category.competition_id,
                        category_id=category.id,
                        participant1_id=p1.id,
                        bracket_type="bye",
                        round_number=round_num,
                        status="bye",
                        winner_id=p1.id
                    )
                    db.session.add(match)
            
            participants_copy = participants_copy[:len(participants_copy)//2] if len(participants_copy) > 2 else []
            round_num += 1
    
    db.session.commit()


# Routes
@app.route("/", methods=["GET"])
def index():
    competitions = Competition.query.order_by(Competition.created_at.desc()).all()
    return render_template("index.html", competitions=competitions)


@app.route("/competition/new", methods=["GET", "POST"])
def new_competition():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        event_date_str = request.form.get("event_date", "")
        
        if not name:
            flash("El nombre del campeonato es obligatorio.")
            return redirect(url_for("new_competition"))
        
        try:
            event_date = datetime.strptime(event_date_str, "%Y-%m-%d").date()
        except:
            event_date = datetime.today().date()
        
        competition = Competition(name=name, event_date=event_date)
        db.session.add(competition)
        db.session.commit()
        
        flash(f"Campeonato '{name}' creado.")
        return redirect(url_for("competition_detail", comp_id=competition.id))
    
    return render_template("new_competition.html")


@app.route("/competition/<int:comp_id>", methods=["GET"])
def competition_detail(comp_id):
    competition = Competition.query.get_or_404(comp_id)
    participant_search = request.args.get("participant_search", "").strip()
    category_search = request.args.get("category_search", "").strip()
    active_tab = request.args.get("active_tab", "")

    participant_query = Participant.query.filter_by(competition_id=comp_id)
    if participant_search:
        # Búsqueda insensible a acentos y mayúsculas usando unidecode
        search_normalized = unidecode(participant_search).lower()
        participant_query = participant_query.filter(
            db.func.lower(Participant.name).like(f"%{search_normalized}%")
        )
    participants = participant_query.order_by(Participant.name).all()
    
    # Si hay búsqueda, filtrar adicionalmente por acentos en Python
    if participant_search:
        search_normalized = unidecode(participant_search).lower()
        participants = [p for p in participants if search_normalized in unidecode(p.name).lower()]
    
    kata_query = Category.query.filter_by(competition_id=comp_id, modality="kata")
    kumite_query = Category.query.filter_by(competition_id=comp_id, modality="kumite")
    if category_search:
        # Búsqueda insensible a acentos y mayúsculas usando unidecode
        search_normalized = unidecode(category_search).lower()
        kata_query = kata_query.filter(
            db.func.lower(Category.name).like(f"%{search_normalized}%")
        )
        kumite_query = kumite_query.filter(
            db.func.lower(Category.name).like(f"%{search_normalized}%")
        )
    
    kata_cats = kata_query.all()
    kumite_cats = kumite_query.all()
    
    # Si hay búsqueda de categorías, filtrar adicionalmente por acentos en Python
    if category_search:
        search_normalized = unidecode(category_search).lower()
        kata_cats = [c for c in kata_cats if search_normalized in unidecode(c.name).lower()]
        kumite_cats = [c for c in kumite_cats if search_normalized in unidecode(c.name).lower()]
    
    return render_template(
        "competition.html",
        competition=competition,
        kata_categories=kata_cats,
        kumite_categories=kumite_cats,
        participants=participants,
        participant_search=participant_search,
        category_search=category_search,
        active_tab=active_tab
    )


@app.route("/competition/<int:comp_id>/settings", methods=["GET", "POST"])
def competition_settings(comp_id):
    competition = Competition.query.get_or_404(comp_id)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        event_date_str = request.form.get("event_date", "")
        
        if name:
            competition.name = name
        
        if event_date_str:
            try:
                competition.event_date = datetime.strptime(event_date_str, "%Y-%m-%d").date()
            except:
                pass
        
        db.session.commit()
        flash("Configuración actualizada.")
        return redirect(url_for("competition_detail", comp_id=comp_id))
    
    return render_template("competition_settings.html", competition=competition)


@app.route("/competition/<int:comp_id>/category/new", methods=["GET", "POST"])
def new_category(comp_id):
    competition = Competition.query.get_or_404(comp_id)
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        modality = request.form.get("modality", "kata")
        group = request.form.get("group", "").strip() or None
        
        if not name:
            flash("El nombre de la categoría es obligatorio.")
            return redirect(url_for("new_category", comp_id=comp_id))
        
        criteria = {}
        
        if modality == "kata":
            age_min = request.form.get("age_min")
            age_max = request.form.get("age_max")
            gender = request.form.get("gender", "X")
            grade_type = request.form.get("grade_type")  # "desde" o "hasta"
            grade_value = request.form.get("grade_value")
            
            if age_min:
                criteria["age_min"] = int(age_min)
            if age_max:
                criteria["age_max"] = int(age_max)
            criteria["gender"] = gender
            
            # Convert grade to numeric range
            if grade_value and grade_type:
                grade_numeric = grade_to_numeric(grade_value)
                if grade_numeric is not None:
                    if grade_type == "desde":
                        # desde 7 = include 7 and higher (stronger grades)
                        criteria["grade_min"] = grade_numeric
                        criteria["grade_min_display"] = grade_value
                    elif grade_type == "hasta":
                        # hasta 7 = include 7 and lower (weaker grades)
                        criteria["grade_max"] = grade_numeric
                        criteria["grade_max_display"] = grade_value
        
        else:  # kumite
            age_min = request.form.get("age_min")
            age_max = request.form.get("age_max")
            gender = request.form.get("gender", "X")
            weight_min = request.form.get("weight_min")
            weight_max = request.form.get("weight_max")
            
            if age_min:
                criteria["age_min"] = int(age_min)
            if age_max:
                criteria["age_max"] = int(age_max)
            criteria["gender"] = gender
            if weight_min:
                criteria["weight_min"] = float(weight_min)
            if weight_max:
                criteria["weight_max"] = float(weight_max)
        
        category = Category(
            competition_id=comp_id,
            name=name,
            modality=modality,
            group=group,
            criteria=criteria
        )
        db.session.add(category)
        db.session.commit()
        
        flash(f"Categoría '{name}' creada.")
        return redirect(url_for("competition_detail", comp_id=comp_id))
    
    return render_template("new_category.html", competition=competition)


@app.route("/category/<int:cat_id>/edit", methods=["GET", "POST"])
def edit_category(cat_id):
    category = Category.query.get_or_404(cat_id)
    competition = category.competition
    
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        modality = request.form.get("modality", category.modality)
        group = request.form.get("group", "").strip() or None
        
        if not name:
            flash("El nombre de la categoría es obligatorio.")
            return redirect(url_for("edit_category", cat_id=cat_id))
        
        criteria = {}
        
        if modality == "kata":
            age_min = request.form.get("age_min")
            age_max = request.form.get("age_max")
            gender = request.form.get("gender", "X")
            grade_type = request.form.get("grade_type")
            grade_value = request.form.get("grade_value")
            
            if age_min:
                criteria["age_min"] = int(age_min)
            if age_max:
                criteria["age_max"] = int(age_max)
            criteria["gender"] = gender
            
            if grade_value and grade_type:
                grade_numeric = grade_to_numeric(grade_value)
                if grade_numeric is not None:
                    if grade_type == "desde":
                        criteria["grade_min"] = grade_numeric
                        criteria["grade_min_display"] = grade_value
                    elif grade_type == "hasta":
                        criteria["grade_max"] = grade_numeric
                        criteria["grade_max_display"] = grade_value
        
        else:  # kumite
            age_min = request.form.get("age_min")
            age_max = request.form.get("age_max")
            gender = request.form.get("gender", "X")
            weight_min = request.form.get("weight_min")
            weight_max = request.form.get("weight_max")
            
            if age_min:
                criteria["age_min"] = int(age_min)
            if age_max:
                criteria["age_max"] = int(age_max)
            criteria["gender"] = gender
            if weight_min:
                criteria["weight_min"] = float(weight_min)
            if weight_max:
                criteria["weight_max"] = float(weight_max)
        
        category.name = name
        category.modality = modality
        category.group = group
        category.criteria = criteria
        db.session.commit()
        
        flash(f"Categoría '{name}' actualizada.")
        return redirect(url_for("competition_detail", comp_id=competition.id))
    
    return render_template("edit_category.html", category=category, competition=competition)


@app.route("/category/<int:cat_id>/delete", methods=["POST"])
def delete_category(cat_id):
    category = Category.query.get_or_404(cat_id)
    competition = category.competition
    
    # Don't allow deleting system categories
    if category.name in ["faltan datos", "extras"]:
        flash("No se pueden eliminar las categorías del sistema.")
        return redirect(url_for("competition_detail", comp_id=competition.id))
    
    extras_category = Category.query.filter_by(
        competition_id=competition.id,
        name="extras",
        modality=category.modality
    ).first()
    
    for participant in list(category.participants):
        # Remove the deleted category assignment for this modality
        for link in list(participant.category_links):
            if link.category_id == category.id:
                db.session.delete(link)
        if extras_category:
            db.session.add(ParticipantCategory(participant=participant, category=extras_category))
    
    db.session.delete(category)
    db.session.commit()
    
    flash(f"Categoría '{category.name}' eliminada.")
    return redirect(url_for("competition_detail", comp_id=competition.id))


@app.route("/competition/<int:comp_id>/upload", methods=["POST"])
def upload_participants(comp_id):
    competition = Competition.query.get_or_404(comp_id)
    
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Selecciona un archivo Excel.")
        return redirect(url_for("competition_detail", comp_id=comp_id))
    
    success, msg = load_participants_from_excel(file, competition)
    flash(msg)
    
    return redirect(url_for("competition_detail", comp_id=comp_id))


@app.route("/competition/<int:comp_id>/assign/<modality>", methods=["POST"])
def assign_and_bracket(comp_id, modality):
    competition = Competition.query.get_or_404(comp_id)
    
    assign_participants_to_categories(competition, modality)
    flash(f"Participantes asignados a categorías de {modality}.")
    
    # Optionally generate brackets
    if request.form.get("generate_brackets"):
        categories = Category.query.filter_by(competition_id=comp_id, modality=modality).all()
        for category in categories:
            generate_brackets(category)
        flash(f"Llaves generadas para {modality}.")
    
    return redirect(url_for("competition_detail", comp_id=comp_id))


@app.route("/category/<int:cat_id>", methods=["GET"])
def category_detail(cat_id):
    category = Category.query.get_or_404(cat_id)
    participants = category.participants
    matches = Match.query.filter_by(category_id=cat_id).order_by(Match.round_number, Match.id).all()
    
    return render_template(
        "category.html",
        category=category,
        participants=participants,
        matches=matches
    )


@app.route("/match/<int:match_id>/result", methods=["POST"])
def record_result(match_id):
    match = Match.query.get_or_404(match_id)
    winner_id = request.form.get("winner_id")
    
    if winner_id:
        match.winner_id = int(winner_id)
        match.status = "completed"
        db.session.commit()
        flash("Resultado registrado.")
    
    return redirect(url_for("category_detail", cat_id=match.category_id))


@app.route("/health")
def health():
    return "OK", 200


@app.route("/competition/<int:comp_id>/participants", methods=["GET"])
def list_participants(comp_id):
    competition = Competition.query.get_or_404(comp_id)
    participants = Participant.query.filter_by(competition_id=comp_id).order_by(Participant.name).all()
    return render_template("participants.html", competition=competition, participants=participants)


@app.route("/participant/<int:part_id>/edit", methods=["GET", "POST"])
def edit_participant(part_id):
    participant = Participant.query.get_or_404(part_id)
    competition = participant.competition
    
    if request.method == "POST":
        participant.name = request.form["name"]
        participant.birthdate = datetime.strptime(request.form["birthdate"], "%Y-%m-%d").date()
        participant.gender = request.form["gender"]
        participant.grade = request.form.get("grade") or None
        participant.weight = float(request.form.get("weight")) if request.form.get("weight") else None
        participant.kata_participation = "kata" in request.form
        participant.kumite_participation = "kumite" in request.form
        db.session.commit()
        flash("Participante actualizado.")
        return redirect(url_for("competition_detail", comp_id=competition.id, participant_search=""))
    
    return render_template("edit_participant.html", participant=participant, competition=competition)


@app.route("/participant/<int:part_id>/delete", methods=["POST"])
def delete_participant(part_id):
    participant = Participant.query.get_or_404(part_id)
    competition = participant.competition
    db.session.delete(participant)
    db.session.commit()
    flash("Participante eliminado.")
    return redirect(url_for("competition_detail", comp_id=competition.id, participant_search=""))


@app.route("/competition/<int:comp_id>/participants/delete_all", methods=["POST"])
def delete_all_participants(comp_id):
    competition = Competition.query.get_or_404(comp_id)
    participants = Participant.query.filter_by(competition_id=comp_id).all()
    for participant in participants:
        db.session.delete(participant)
    db.session.commit()
    flash("Todos los participantes han sido eliminados.")
    return redirect(url_for("list_participants", comp_id=comp_id))


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug_mode, host="0.0.0.0", port=port)

